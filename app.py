import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from scipy.optimize import linprog
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import base64
from fpdf import FPDF
import openpyxl
from openpyxl import Workbook
import tempfile
import os

# Configuração da página
st.set_page_config(
    page_title="Solver de Programação Linear - Lachtermacher 3.1.1",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Título principal
st.title("🎯 Solver de Programação Linear")
st.subheader("Problema 3.1.1 - Lachtermacher: Pesquisa Operacional na Tomada de Decisões")
st.markdown("---")

# Função para resolver o problema de PL
def solve_lp_problem(c, A, b, sense, objective_type):
    """
    Resolve problema de programação linear usando scipy.optimize.linprog
    
    Args:
        c: coeficientes da função objetivo
        A: matriz de restrições
        b: lado direito das restrições
        sense: sentido das desigualdades ('<=', '>=', '=')
        objective_type: 'max' ou 'min'
    """
    try:
        # Converter para minimização (linprog sempre minimiza)
        if objective_type == 'max':
            c = [-coef for coef in c]
        
        # Converter restrições para formato padrão do linprog
        A_eq = []
        b_eq = []
        A_ub = []
        b_ub = []
        
        for i, constraint_sense in enumerate(sense):
            if constraint_sense == '=':
                A_eq.append(A[i])
                b_eq.append(b[i])
            elif constraint_sense == '<=':
                A_ub.append(A[i])
                b_ub.append(b[i])
            elif constraint_sense == '>=':
                A_ub.append([-a for a in A[i]])
                b_ub.append(-b[i])
        
        # Converter para arrays numpy
        A_eq = np.array(A_eq) if A_eq else None
        b_eq = np.array(b_eq) if b_eq else None
        A_ub = np.array(A_ub) if A_ub else None
        b_ub = np.array(b_ub) if b_ub else None
        
        # Resolver o problema
        result = linprog(
            c=c,
            A_ub=A_ub,
            b_ub=b_ub,
            A_eq=A_eq,
            b_eq=b_eq,
            bounds=[(0, None), (0, None)],  # x1, x2 >= 0
            method='highs'
        )
        
        return result
        
    except Exception as e:
        st.error(f"Erro ao resolver o problema: {str(e)}")
        return None

# Função para gerar gráfico da região viável
def plot_feasible_region(A, b, sense, optimal_point=None, objective_coeffs=None):
    """
    Gera gráfico da região viável usando Plotly
    """
    # Criar grade de pontos
    x1_range = np.linspace(0, 20, 1000)
    x2_range = np.linspace(0, 20, 1000)
    X1, X2 = np.meshgrid(x1_range, x2_range)
    
    # Verificar viabilidade para cada ponto
    feasible = np.ones_like(X1, dtype=bool)
    
    for i, (constraint, rhs, constraint_sense) in enumerate(zip(A, b, sense)):
        if constraint_sense == '<=':
            feasible &= (constraint[0] * X1 + constraint[1] * X2 <= rhs)
        elif constraint_sense == '>=':
            feasible &= (constraint[0] * X1 + constraint[1] * X2 >= rhs)
        elif constraint_sense == '=':
            feasible &= (np.abs(constraint[0] * X1 + constraint[1] * X2 - rhs) < 0.1)
    
    # Criar figura Plotly
    fig = go.Figure()
    
    # Adicionar região viável
    if np.any(feasible):
        fig.add_trace(go.Scatter(
            x=X1[feasible],
            y=X2[feasible],
            mode='markers',
            marker=dict(size=2, color='lightblue', opacity=0.6),
            name='Região Viável',
            showlegend=True
        ))
    
    # Adicionar linhas de restrições
    colors = ['red', 'blue', 'green', 'orange', 'purple', 'brown']
    for i, (constraint, rhs, constraint_sense) in enumerate(zip(A, b, sense)):
        color = colors[i % len(colors)]
        
        # Calcular pontos para a linha de restrição
        if constraint[1] != 0:  # Se coeficiente de x2 não é zero
            x1_points = np.linspace(0, 20, 100)
            x2_points = (rhs - constraint[0] * x1_points) / constraint[1]
        else:  # Linha vertical
            x1_points = np.full(100, rhs / constraint[0])
            x2_points = np.linspace(0, 20, 100)
        
        # Filtrar pontos válidos
        valid_mask = (x1_points >= 0) & (x2_points >= 0) & (x1_points <= 20) & (x2_points <= 20)
        
        if np.any(valid_mask):
            fig.add_trace(go.Scatter(
                x=x1_points[valid_mask],
                y=x2_points[valid_mask],
                mode='lines',
                line=dict(color=color, width=2),
                name=f'Restrição {i+1}: {constraint[0]}x₁ + {constraint[1]}x₂ {constraint_sense} {rhs}',
                showlegend=True
            ))
    
    # Adicionar ponto ótimo se disponível
    if optimal_point is not None:
        fig.add_trace(go.Scatter(
            x=[optimal_point[0]],
            y=[optimal_point[1]],
            mode='markers',
            marker=dict(size=12, color='gold', symbol='star'),
            name='Solução Ótima',
            showlegend=True
        ))
    
    # Adicionar linha de isoprofit se disponível
    if objective_coeffs is not None and optimal_point is not None:
        # Calcular valor da função objetivo no ponto ótimo
        optimal_value = objective_coeffs[0] * optimal_point[0] + objective_coeffs[1] * optimal_point[1]
        
        # Gerar linha de isoprofit
        x1_isoprofit = np.linspace(0, 20, 100)
        if objective_coeffs[1] != 0:
            x2_isoprofit = (optimal_value - objective_coeffs[0] * x1_isoprofit) / objective_coeffs[1]
        else:
            x2_isoprofit = np.linspace(0, 20, 100)
        
        valid_mask = (x1_isoprofit >= 0) & (x2_isoprofit >= 0) & (x1_isoprofit <= 20) & (x2_isoprofit <= 20)
        
        if np.any(valid_mask):
            fig.add_trace(go.Scatter(
                x=x1_isoprofit[valid_mask],
                y=x2_isoprofit[valid_mask],
                mode='lines',
                line=dict(color='black', width=2, dash='dash'),
                name=f'Isoprofit: {objective_coeffs[0]}x₁ + {objective_coeffs[1]}x₂ = {optimal_value:.2f}',
                showlegend=True
            ))
    
    # Configurar layout
    fig.update_layout(
        title="Região Viável e Solução Ótima",
        xaxis_title="x₁",
        yaxis_title="x₂",
        xaxis=dict(range=[0, 20]),
        yaxis=dict(range=[0, 20]),
        width=800,
        height=600,
        showlegend=True,
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1
        )
    )
    
    return fig

# Função para exportar resultados em CSV
def export_to_csv(result_data, optimal_point, objective_value, objective_coeffs):
    """Exporta resultados para CSV"""
    df_results = pd.DataFrame({
        'Variável': ['x₁', 'x₂'],
        'Valor Ótimo': [optimal_point[0], optimal_point[1]]
    })
    
    df_objective = pd.DataFrame({
        'Função Objetivo': [f'{objective_coeffs[0]}x₁ + {objective_coeffs[1]}x₂'],
        'Valor Ótimo': [objective_value]
    })
    
    # Criar buffer para CSV
    output = io.StringIO()
    output.write("RESULTADOS DA PROGRAMAÇÃO LINEAR\n")
    output.write("=" * 40 + "\n\n")
    
    df_objective.to_csv(output, index=False)
    output.write("\n")
    df_results.to_csv(output, index=False)
    
    return output.getvalue()

# Função para exportar resultados em Excel
def export_to_excel(result_data, optimal_point, objective_value, objective_coeffs, fig):
    """Exporta resultados para Excel"""
    wb = Workbook()
    
    # Planilha de resultados
    ws1 = wb.active
    ws1.title = "Resultados"
    
    ws1['A1'] = "RESULTADOS DA PROGRAMAÇÃO LINEAR"
    ws1['A2'] = "=" * 40
    
    ws1['A4'] = "Função Objetivo"
    ws1['B4'] = f'{objective_coeffs[0]}x₁ + {objective_coeffs[1]}x₂'
    ws1['A5'] = "Valor Ótimo"
    ws1['B5'] = objective_value
    
    ws1['A7'] = "Variável"
    ws1['B7'] = "Valor Ótimo"
    ws1['A8'] = "x₁"
    ws1['B8'] = optimal_point[0]
    ws1['A9'] = "x₂"
    ws1['B9'] = optimal_point[1]
    
    # Adicionar informações sobre o gráfico (sem a imagem)
    ws2 = wb.create_sheet("Informações do Gráfico")
    ws2['A1'] = "INFORMAÇÕES DO GRÁFICO"
    ws2['A3'] = "O gráfico da região viável foi gerado com sucesso."
    ws2['A4'] = "Para visualizar o gráfico, use a interface do Streamlit."
    ws2['A6'] = "Coordenadas do ponto ótimo:"
    ws2['A7'] = f"x₁ = {optimal_point[0]:.4f}"
    ws2['A8'] = f"x₂ = {optimal_point[1]:.4f}"
    
    # Salvar para buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# Função para exportar resultados em PDF
def export_to_pdf(result_data, optimal_point, objective_value, objective_coeffs):
    """Exporta resultados para PDF"""
    pdf = FPDF()
    pdf.add_page()
    
    # Título
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Resultados da Programacao Linear', ln=True, align='C')
    pdf.ln(10)
    
    # Função objetivo
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, 'Funcao Objetivo:', ln=True)
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, f'{objective_coeffs[0]}x1 + {objective_coeffs[1]}x2', ln=True)
    pdf.cell(0, 10, f'Valor Otimo: {objective_value:.4f}', ln=True)
    pdf.ln(10)
    
    # Valores das variáveis
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 10, 'Valores Otimos das Variaveis:', ln=True)
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, f'x1 = {optimal_point[0]:.4f}', ln=True)
    pdf.cell(0, 10, f'x2 = {optimal_point[1]:.4f}', ln=True)
    
    # Salvar para arquivo temporário e depois ler como bytes
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
        pdf.output(tmp_file.name)
        
        # Ler o arquivo como bytes
        with open(tmp_file.name, 'rb') as f:
            pdf_bytes = f.read()
        
        # Limpar arquivo temporário
        os.unlink(tmp_file.name)
    
    return pdf_bytes

# Sidebar para entrada de dados
with st.sidebar:
    st.header("📝 Definição do Problema")
    
    # Tipo de objetivo
    objective_type = st.selectbox(
        "Tipo de Objetivo:",
        ["max", "min"],
        format_func=lambda x: "Maximizar" if x == "max" else "Minimizar"
    )
    
    # Coeficientes da função objetivo
    st.subheader("Função Objetivo")
    c1 = st.number_input("Coeficiente de x₁:", value=3.0, step=0.1)
    c2 = st.number_input("Coeficiente de x₂:", value=2.0, step=0.1)
    
    # Número de restrições
    num_constraints = st.slider("Número de Restrições:", 1, 6, 4)
    
    # Coeficientes das restrições
    st.subheader("Restrições")
    A = []
    b = []
    sense = []
    restricoes_padrao = [
        {'a1': 1.0, 'a2': 2.0, 'sense': '<=', 'rhs': 6.0},
        {'a1': 2.0, 'a2': 1.0, 'sense': '<=', 'rhs': 8.0},
        {'a1': -1.0, 'a2': 1.0, 'sense': '<=', 'rhs': 1.0},
        {'a1': 0.0, 'a2': 1.0, 'sense': '<=', 'rhs': 2.0},
    ]
    for i in range(num_constraints):
        st.markdown(f"**Restrição {i+1}:**")
        col1, col2, col3, col4 = st.columns([2, 2, 2, 1])
        padrao = restricoes_padrao[i] if i < len(restricoes_padrao) else {'a1': 1.0, 'a2': 1.0, 'sense': '<=', 'rhs': 1.0}
        with col1:
            a1 = st.number_input(f"x₁", value=padrao['a1'], key=f"a1_{i}")
        with col2:
            a2 = st.number_input(f"x₂", value=padrao['a2'], key=f"a2_{i}")
        with col3:
            sense_type = st.selectbox("Tipo", ["<=", ">=", "="], index=["<=", ">=", "="].index(padrao['sense']), key=f"sense_{i}")
        with col4:
            rhs = st.number_input("RHS", value=padrao['rhs'], key=f"rhs_{i}")
        A.append([a1, a2])
        b.append(rhs)
        sense.append(sense_type)
    
    # Botão para resolver
    solve_button = st.button("🔍 Resolver Problema", type="primary")

# Área principal
if solve_button:
    st.header("📊 Resultados")
    
    # Resolver o problema
    result = solve_lp_problem([c1, c2], A, b, sense, objective_type)
    
    if result is not None:
        # Exibir resultados
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Solução Ótima")
            
            if result.success:
                # Calcular valor ótimo (considerar se é maximização ou minimização)
                if objective_type == 'max':
                    optimal_value = -result.fun
                else:
                    optimal_value = result.fun
                
                optimal_point = result.x
                
                st.success("✅ Problema resolvido com sucesso!")
                st.metric("Valor Ótimo da Função Objetivo", f"{optimal_value:.4f}")
                st.metric("x₁ (ótimo)", f"{optimal_point[0]:.4f}")
                st.metric("x₂ (ótimo)", f"{optimal_point[1]:.4f}")
                
                # Status da solução
                if result.status == 0:
                    st.info("Status: Solução ótima encontrada")
                elif result.status == 1:
                    st.warning("Status: Iteração limitada atingida")
                elif result.status == 2:
                    st.error("Status: Problema inviável")
                elif result.status == 3:
                    st.error("Status: Problema ilimitado")
                else:
                    st.warning(f"Status: {result.status}")
                
            else:
                st.error("❌ Problema não pôde ser resolvido")
                st.write(f"Status: {result.message}")
        
        with col2:
            st.subheader("Informações do Problema")
            st.write(f"**Função Objetivo:** {'Maximizar' if objective_type == 'max' else 'Minimizar'} {c1}x₁ + {c2}x₂")
            st.write("**Restrições:**")
            for i, (constraint, rhs, constraint_sense) in enumerate(zip(A, b, sense)):
                st.write(f"{i+1}. {constraint[0]}x₁ + {constraint[1]}x₂ {constraint_sense} {rhs}")
        
        # Gráfico da região viável
        if result.success:
            st.header("📈 Gráfico da Região Viável")
            
            # Calcular coeficientes corretos para o gráfico
            if objective_type == 'max':
                plot_objective_coeffs = [c1, c2]
            else:
                plot_objective_coeffs = [c1, c2]
            
            fig = plot_feasible_region(A, b, sense, optimal_point, plot_objective_coeffs)
            st.plotly_chart(fig, use_container_width=True)
            
            # Exportação de resultados
            st.header("💾 Exportação de Resultados")
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                # CSV
                csv_data = export_to_csv(None, optimal_point, optimal_value, [c1, c2])
                st.download_button(
                    label="📄 Download CSV",
                    data=csv_data,
                    file_name="resultados_pl.csv",
                    mime="text/csv"
                )
            
            with col2:
                # Excel
                excel_data = export_to_excel(None, optimal_point, optimal_value, [c1, c2], fig)
                st.download_button(
                    label="📊 Download Excel",
                    data=excel_data,
                    file_name="resultados_pl.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            with col3:
                # PDF
                pdf_data = export_to_pdf(None, optimal_point, optimal_value, [c1, c2])
                st.download_button(
                    label="📋 Download PDF",
                    data=pdf_data,
                    file_name="resultados_pl.pdf",
                    mime="application/pdf"
                )
            
            with col4:
                # PNG do gráfico (usando matplotlib como alternativa)
                try:
                    # Tentar usar Plotly com Kaleido
                    img_bytes = fig.to_image(format="png")
                    st.download_button(
                        label="🖼️ Download PNG",
                        data=img_bytes,
                        file_name="grafico_regiao_viavel.png",
                        mime="image/png"
                    )
                except Exception as e:
                    # Se falhar, mostrar mensagem informativa
                    st.info("🖼️ Download PNG: Use a função de captura de tela do navegador para salvar o gráfico")
                    st.caption("(Chrome não encontrado para exportação automática)")

# Informações adicionais
else:
    st.info("👈 Use a barra lateral para definir o problema de programação linear e clique em 'Resolver Problema'")
    
    # Exemplo do problema do Excel Solver
    st.header("📚 Exemplo: Problema de Programação Linear (igual ao Excel Solver)")
    st.markdown("""
    **Problema:** Maximizar Z = 3x₁ + 2x₂
    
    **Sujeito a:**
    - x₁ + 2x₂ ≤ 6
    - 2x₁ + x₂ ≤ 8
    - -x₁ + x₂ ≤ 1
    - x₂ ≤ 2
    - x₁, x₂ ≥ 0
    
    **Solução:** x₁ = 3,333..., x₂ = 1,333..., Z = 12,666...
    """)
    
    # Instruções de uso
    st.header("🎯 Como Usar")
    st.markdown("""
    1. **Defina o tipo de objetivo** (Maximizar ou Minimizar)
    2. **Insira os coeficientes** da função objetivo
    3. **Configure as restrições** (coeficientes, tipo de desigualdade e RHS)
    4. **Clique em 'Resolver Problema'**
    5. **Visualize os resultados** e o gráfico da região viável
    6. **Exporte os resultados** nos formatos desejados
    """)

# Footer
st.markdown("---")
st.markdown("**Desenvolvido para o Problema 3.1.1 do livro 'Pesquisa Operacional na Tomada de Decisões' de Gerson Lachtermacher**") 